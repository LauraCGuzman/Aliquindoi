"""
Generador de datos sintéticos para demo de Aliquindoi.
Genera archivos .asc (espectrofotómetro UV) y .xlsx (FTIR) con datos realistas.

Estructura generada:
  demo_data/
    UV/
      sample_SampleA-1.Sample.asc
      sample_SampleA-2.Sample.asc
      sample_SampleB-1.Sample.asc
      sample_SampleB-2.Sample.asc
      zero_demo.asc
      base_demo.asc
    FTIR/
      20260409data.xlsx
      20260409ref.xlsx
"""

import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ── rutas ──────────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UV_DIR  = os.path.join(BASE_DIR, "UV")
IR_DIR  = os.path.join(BASE_DIR, "FTIR")
os.makedirs(UV_DIR,  exist_ok=True)
os.makedirs(IR_DIR, exist_ok=True)

FECHA_STR   = "26/04/09"   # yy/mm/dd  (para cabeceras .asc)
HORA_BASE   = "09:00:00.000000"
HORA_MUESTRA_A1 = "09:30:00.000000"
HORA_MUESTRA_A2 = "09:35:00.000000"
HORA_MUESTRA_B1 = "10:00:00.000000"
HORA_MUESTRA_B2 = "10:05:00.000000"
FECHA_ID    = "20260409"   # yyyyMMdd  (para nombres de archivos FTIR)

np.random.seed(42)

# ── rangos espectrales ─────────────────────────────────────────────────────────
NM_UV  = np.arange(250, 2505, 5)          # 250-2500 nm (espectrofotómetro)
NM_IR  = np.arange(2500, 16050, 50, dtype=float)  # 2500-16000 nm (FTIR, en nm)

# ═══════════════════════════════════════════════════════════════════════════════
#  Funciones de perfil espectral sintético
# ═══════════════════════════════════════════════════════════════════════════════

def perfil_absorber_uv(nm, ruido=0.005):
    """
    Perfil de reflectancia de un absorbedor solar selectivo (tipo cermet).
    Alta absortancia solar (baja reflectancia en UV-VIS-NIR) → reflectancia ~5-15 %.
    """
    base = 0.05 + 0.06 * np.exp(-(nm - 1800) ** 2 / (2 * 600 ** 2))
    base += np.random.normal(0, ruido, size=len(nm))
    return np.clip(base, 0.01, 0.98)


def perfil_reflector_uv(nm, ruido=0.005):
    """
    Perfil de reflectancia de un reflector solar (tipo aluminio pulido).
    Alta reflectancia en todo el visible-NIR → reflectancia ~85-93 %.
    """
    base = 0.88 - 0.04 * np.exp(-(nm - 300) ** 2 / (2 * 100 ** 2))
    base += 0.02 * np.sin(nm / 200)
    base += np.random.normal(0, ruido, size=len(nm))
    return np.clip(base, 0.01, 0.99)


def perfil_zero(nm):
    """Zero line: señal cercana a cero."""
    return np.random.normal(0.001, 0.0005, size=len(nm))


def perfil_base(nm, escala=1.0):
    """Baseline de referencia: señal plana alta (patrón blanco)."""
    base = escala * (0.99 + np.random.normal(0, 0.002, size=len(nm)))
    return np.clip(base, 0.85, 1.02)


# ═══════════════════════════════════════════════════════════════════════════════
#  Generador de archivos .asc
# ═══════════════════════════════════════════════════════════════════════════════

def escribir_asc(ruta, nm_array, valores, fecha, hora, titulo="Demo"):
    """
    Escribe un archivo .asc compatible con el espectrofotómetro Perkin-Elmer / Shimadzu.

    Formato esperado por el programa:
      línea 0 : nombre
      línea 1 : vacío
      línea 2 : vacío
      línea 3 : fecha  yy/mm/dd
      línea 4 : hora   HH:MM:SS.ffffff
      líneas 5+: cabecera libre
      #DATA
      nm\tvalor   (coma como separador decimal)
    """
    with open(ruta, "w", encoding="utf-8") as f:
        f.write(f"{titulo}\n")
        f.write("\n")
        f.write("\n")
        f.write(f"{fecha}\n")
        f.write(f"{hora}\n")
        f.write("Unidades: nm / %R\n")
        f.write("Instrumento: Demo Spectrophotometer\n")
        f.write("#DATA\n")
        for nm_val, val in zip(nm_array, valores):
            # coma como separador decimal, tabulador entre columnas
            f.write(f"{str(nm_val).replace('.', ',')}\t{str(round(float(val), 6)).replace('.', ',')}\n")
    print(f"  OK  {os.path.basename(ruta)}")


# ═══════════════════════════════════════════════════════════════════════════════
#  Generador de archivos FTIR .xlsx
# ═══════════════════════════════════════════════════════════════════════════════

def perfil_absorber_ir(nm, ruido=0.005):
    """
    Perfil de reflectancia IR de un absorbedor solar (~5-20 % de reflectancia
    en el rango térmico, equivale a emitancia ~80-95 %).
    """
    base = 0.08 + 0.08 * np.sin((nm - 2500) / 5000)
    base += np.random.normal(0, ruido, size=len(nm))
    return np.clip(base, 0.01, 0.99)


def perfil_reflector_ir(nm, ruido=0.005):
    """
    Perfil de reflectancia IR de un reflector (~90-95 % en IR).
    """
    base = 0.92 - 0.02 * np.exp(-(nm - 10000) ** 2 / (2 * 4000 ** 2))
    base += np.random.normal(0, ruido, size=len(nm))
    return np.clip(base, 0.01, 0.99)


def perfil_zero_ir(nm):
    return np.random.normal(0.002, 0.001, size=len(nm))


def perfil_base_ir(nm):
    base = 0.98 + np.random.normal(0, 0.002, size=len(nm))
    return np.clip(base, 0.90, 1.02)


def escribir_hoja_ftir(ws, nm_array, valores, nombre_hoja):
    """
    Escribe una hoja FTIR con el formato que espera el programa:
      - 4 filas de cabecera (índices 0-3) → header=4 en pandas
      - Fila 4 (índice 4) = cabecera de columnas: nm  %R
      - A partir de fila 5: datos
    """
    ws.title = nombre_hoja
    # 4 filas de cabecera libre
    ws.append(["Instrumento: Demo FTIR"])
    ws.append(["Muestra: " + nombre_hoja])
    ws.append(["Unidades: nm / %R"])
    ws.append([])
    # Fila de encabezado (será la row index 4, leída por pandas con header=4)
    ws.append(["nm", "%R"])
    # Datos
    for nm_val, val in zip(nm_array, valores):
        ws.append([float(nm_val), round(float(val * 100), 4)])  # %R → multiplicado por 100


def crear_excel_ftir_data(ruta, muestras_dict):
    """
    Crea el archivo Excel de datos FTIR.
    muestras_dict = { nombre_hoja: array_reflectancia }
    """
    wb = Workbook()
    primera = True
    for nombre_hoja, valores in muestras_dict.items():
        if primera:
            ws = wb.active
            primera = False
        else:
            ws = wb.create_sheet()
        escribir_hoja_ftir(ws, NM_IR, valores, nombre_hoja)
    wb.save(ruta)
    print(f"  OK  {os.path.basename(ruta)}")


def crear_excel_ftir_ref(ruta, refs_dict):
    """
    Crea el archivo Excel de referencias FTIR.
    refs_dict = { nombre_hoja: array_reflectancia }
    """
    wb = Workbook()
    primera = True
    for nombre_hoja, valores in refs_dict.items():
        if primera:
            ws = wb.active
            primera = False
        else:
            ws = wb.create_sheet()
        escribir_hoja_ftir(ws, NM_IR, valores, nombre_hoja)
    wb.save(ruta)
    print(f"  OK  {os.path.basename(ruta)}")


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    print("\n=== Generando datos UV (espectrofotómetro) ===")

    # Zero y base UV
    escribir_asc(
        os.path.join(UV_DIR, "zero_demo.asc"),
        NM_UV, perfil_zero(NM_UV),
        FECHA_STR, HORA_BASE, titulo="ZeroLine")

    escribir_asc(
        os.path.join(UV_DIR, "base_demo.asc"),
        NM_UV, perfil_base(NM_UV),
        FECHA_STR, HORA_BASE, titulo="BaseLine")

    # SampleA (absorbedor): 2 réplicas
    for i, hora in enumerate([HORA_MUESTRA_A1, HORA_MUESTRA_A2], start=1):
        escribir_asc(
            os.path.join(UV_DIR, f"sample_SampleA-{i}.Sample.asc"),
            NM_UV, perfil_absorber_uv(NM_UV),
            FECHA_STR, hora, titulo=f"SampleA replica {i}")

    # SampleB (reflector): 2 réplicas
    for i, hora in enumerate([HORA_MUESTRA_B1, HORA_MUESTRA_B2], start=1):
        escribir_asc(
            os.path.join(UV_DIR, f"sample_SampleB-{i}.Sample.asc"),
            NM_UV, perfil_reflector_uv(NM_UV),
            FECHA_STR, hora, titulo=f"SampleB replica {i}")

    print("\n=== Generando datos FTIR ===")

    # Archivo de muestras FTIR
    muestras_ftir = {
        "sample_SampleA-1": perfil_absorber_ir(NM_IR),
        "sample_SampleA-2": perfil_absorber_ir(NM_IR, ruido=0.006),
        "sample_SampleB-1": perfil_reflector_ir(NM_IR),
        "sample_SampleB-2": perfil_reflector_ir(NM_IR, ruido=0.006),
    }
    crear_excel_ftir_data(
        os.path.join(IR_DIR, f"{FECHA_ID}data.xlsx"),
        muestras_ftir)

    # Archivo de referencias FTIR
    refs_ftir = {
        f"zero_{FECHA_ID}": perfil_zero_ir(NM_IR),
        f"base_{FECHA_ID}": perfil_base_ir(NM_IR),
    }
    crear_excel_ftir_ref(
        os.path.join(IR_DIR, f"{FECHA_ID}ref.xlsx"),
        refs_ftir)

    print("\n=== Resumen ===")
    print(f"UV  → {UV_DIR}")
    print(f"FTIR → {IR_DIR}")
    print("\nDone. Para usar con Aliquindoi:")
    print("  - Tipo de medida: Absortancia")
    print("  - Aparatos: FTIR + Espectrofotómetro (sin ventana)")
    print(f"  - Carpeta UV:   {UV_DIR}")
    print(f"  - Carpeta FTIR: {IR_DIR}")


if __name__ == "__main__":
    main()
